from pprint import pprint
import os

import logging
from django.core.management.base import BaseCommand

import openpyxl

from database.models import Table, Column

logger = logging.getLogger(__name__)

TABLE_SHEET = "DBTables"
COLUMN_SHEET = "DBColumns"

class Command(BaseCommand):
    help = "Import data from an Excel file"


    column_cache = {}
    table_cache = {}

    def add_arguments(self, parser):
        parser.add_argument('file', nargs='+')

    def handle(self, *args, **options):
        file = options.get("file")[0]
        print("Reading in file: {}".format(file))
        self.import_data(file=file)

    def import_data(self, file=None):
        workbook = openpyxl.load_workbook(file)
        self.import_table_data(workbook)
        self.import_column_data(workbook)

    def import_table_data(self, workbook=None):
        sheet = workbook.get_sheet_by_name(TABLE_SHEET)
        for row in range(2, sheet.max_row + 1):
            table_name = sheet["A{}".format(row)].value
            if not table_name:
                continue
            use_for_bi = sheet["D{}".format(row)].value == "YES"
            description = sheet["E{}".format(row)].value
            defaults = {
                "use_for_bi": use_for_bi,
                "description": description,
            }
            print("Updating {}".format(table_name))
            Table.objects.update_or_create(name=table_name,
                                           defaults=defaults)

    def import_column_data(self, workbook=None):
        sheet = workbook.get_sheet_by_name(COLUMN_SHEET)
        for row in range(2, sheet.max_row + 1):
            table_name = sheet["A{}".format(row)].value
            if not table_name:
                continue
            column_name = sheet["B{}".format(row)].value
            is_data_structured = sheet["D{}".format(row)].value
            has_duplicates = sheet["E{}".format(row)].value
            needs_index = sheet["F{}".format(row)].value
            table = self._get_table(table_name=table_name)
            defaults = {
                "is_structured": is_data_structured,
                "has_duplicates": has_duplicates,
                "needs_index": needs_index
            }
            print("Updating {}.{}".format(table_name, column_name))
            column, _ = Column.objects.update_or_create(table=table,
                                                        name=column_name,
                                                        defaults=defaults)

    def _get_table(self, table_name=None):
        if table_name not in self.table_cache:
            table = Table.objects.get(name=table_name)
            self.table_cache[table_name] = table
        return self.table_cache.get(table_name)

